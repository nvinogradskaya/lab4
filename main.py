import openpyxl

wb = openpyxl.load_workbook(filename = 'student4.xlsx')
sheet = wb['Sheet1']
wb1 = openpyxl.load_workbook(filename = 'student1.xlsx')
sheet1 = wb1['Sheet1']
ph = []
with open('nom.txt') as f:
    for line in f:
        ph.append(line)

ph1 = {}
for i in range(1000):
    s = ph[i]
    ph1[s[0:40]] = s[41:53]
rus = "абвгдежзийклмнопрстуфхцчшщъыьэюя"
ruszagl = "АБВГДЕЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ"
eng = "abcdefghijklmnopqrstuvwxyz"
for i in range(2, 1001):
    s = ''
    s = sheet.cell(row=i, column=1).value
    for j in ph1:
        if j == s:
            sheet1.cell(row=i, column=1).value = ph1[j]
    s1 = ''
    s = sheet.cell(row = i, column = 3).value
    ind = s.rfind('.') - 2
    raz = rus.find('к') - rus.find(s[ind])
    for j in range(len(s)):
        if s[j] in rus:
            x = (rus.find(s[j]) + raz) % 32
            s1 += rus[x]
        elif s[j] in ruszagl:
            x = (ruszagl.find(s[j]) + raz) % 32
            s1 += ruszagl[x]
        else:
            s1 += s[j]
    sheet1.cell(row=i, column=3).value = s1
    s = ' '
    s1 = ''
    s = sheet.cell(row=i, column=2).value
    for j in range(len(s)):
        if s[j] in eng:
            x = (eng.find(s[j]) + raz) % 26
            s1 += eng[x]
        else:
            s1 += s[j]
    sheet1.cell(row=i, column=2).value = s1
    sheet1.cell(row=i, column=4).value = raz



wb1.save('student1.xlsx')